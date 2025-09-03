import tkinter as tk
import openpyxl
from tkinter import Tk
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side


def remove_columns(column_values):
    """Функция для удаления лишних колонок"""
    for sublist in column_values:
        if len(sublist) > 0:
            sublist.pop(0)
        if len(sublist) > 2:
            sublist.pop(1) 
        if len(sublist) > 0:
            sublist.pop(1)

    return column_values

def create_excel_file(window, fields_list, error_label):
    """Функция для чтения и создания нового excel файла"""
    field_values = []
    column_values = []
    for index, field in enumerate(fields_list):
        value = field.get()
        field_values.append(value)

    read_path = field_values[0]
    column_name = field_values[1]
    value_name = field_values[2]
    file_name = field_values[3]
    write_path = field_values[4]

    try:
        if not read_path.endswith('.xlsx'):
            read_path += '.xlsx'
        workbook = load_workbook(read_path)
    except Exception:
        readpath_error = 'Файл по указанному пути не найден'
        error_label.config(text=f'Ошибка: {readpath_error}')
        return None

    if '' in [column_name, value_name, file_name, write_path]:
        field_error = 'Не все поля заполнены'
        error_label.config(text=f'Ошибка: {field_error}')
        return None
    
    # Чтение и сохранение данных из excel файла указанного пользователем
    sheet = workbook.active
    start_row = 10
    skip_row = 10
    row_for_clear = False
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
    name_columns = []

    for col in columns:
        column_value = sheet[f'{col}{start_row-1}'].value
        name_columns.append(column_value)

    for row in range(21):
        values_row = []
        for index, col in enumerate(columns):
            value = sheet[f'{col}{row+start_row}'].value
            if (row + start_row == skip_row):
                continue
            if (str(name_columns[index]).lower() == column_name.lower() and str(value).lower() == value_name.lower()):
                break
            values_row.append(value)
        if len(values_row) == len(columns):
            column_values.append(values_row)

    column_values.insert(0, name_columns)
    cleared_columns = remove_columns(column_values)

    # Создание и форматирование нового excel файла
    new_workbook = Workbook()
    new_sheet = new_workbook.active

    for row in cleared_columns:
        new_sheet.append(row)

    bold_font = Font(bold=True)
    side = Side(border_style='thin', color='000000')
    border = Border(left=side, right=side, top=side, bottom=side)
    for cell in new_sheet[1]:
        cell.font = bold_font

    for row in new_sheet.iter_rows(min_row=1, max_row=new_sheet.max_row, min_col=1, max_col=new_sheet.max_column):
        for cell in row:
            cell.border = border

    column_widths = [9.84, 11, 8.57, 12.71, 10.86]
    for i, width in enumerate(column_widths, start=1):
        col_letter = new_sheet.cell(row=1, column=i).column_letter
        new_sheet.column_dimensions[col_letter].width = width * 1.1

    if write_path[-1] != '/':
        write_path += '/'
    if not file_name.endswith('.xlsx'):
        file_name += '.xlsx'

    try:
        new_workbook.save(f"{write_path}{file_name}")
    except Exception as e:
        writepath_error = 'Путь для сохранения файла неверный'
        error_label.config(text=f'Ошибка: {writepath_error}')
        return None

    window.destroy()

def create_field_widget(window, fields_list):
    """Функция для создания текстового поля"""
    field = tk.Entry(
        window,
        font=('Arial', 10),
        width=30,
        bg='#ffffff',
    )
    field.pack()
    fields_list.append(field)

def create_label_widget(window, text):
    """Функция для создания текстовой метки"""
    label = tk.Label(
        window,
        text=text,
        font=('Arial', 10, 'bold'),
        fg='#333333',
        bg='#fff3eb'
    )
    label.pack(pady=(10, 0))

def create_button_widget(window, text, fields_list, error_label):
    """Функция для создания кнопки"""
    button = tk.Button(
        window,
        text=text,
        command=lambda: create_excel_file(window, fields_list, error_label),
        font=('Arial', 10, 'bold'),
        fg='#333333',
        bg='#fff3eb'
    )
    button.pack(pady=(40, 0))

def main():
    """Основная функция с созданием графического окна"""
    window = Tk()
    window.title('Тестовый скрипт')
    window.geometry('800x400')
    window.configure(bg='#fff3eb')
    fields_list = []

    # Задаем основные виджеты: Label, Entry, Button
    create_label_widget(window, 'Укажите путь до вашего файла:')
    create_field_widget(window, fields_list)
    create_label_widget(window, 'Укажите название вашего столбца:')
    create_field_widget(window, fields_list)
    create_label_widget(window, 'Укажите название вашего значения:')
    create_field_widget(window, fields_list)
    create_label_widget(window, 'Укажите название вашего нового файла:')
    create_field_widget(window, fields_list)
    create_label_widget(window, 'Укажите путь для сохранения файла:')
    create_field_widget(window, fields_list)
    error_label = tk.Label(window, text="", fg="red", bg='#fff3eb')
    error_label.pack()

    create_button_widget(window, 'Запустить', fields_list, error_label)

    window.mainloop()

if __name__ == "__main__":
    main()